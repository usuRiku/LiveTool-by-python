import openpyxl
import random
import copy
import tqdm
# from tkinter import filedialog


def main():
    # typ = [('Excelファイル', '*.xlsx')]
    # dir = 'C:'
    # tmp_str = input("バンド一覧のExcelファイルを入力してください")
    # fle = filedialog.askopenfilename(filetypes=typ, initialdir=dir)
    print("操作方法は説明書をご覧ください")
    wb_band = WorkBook("band.xlsx")
    # tmp_str = input("今回のライブの出演リストを選択してください")
    wb_ap_band = WorkBook("ap_band.xlsx")
    bands = Bands(wb_band)
    ap_bands = ApBands(wb_ap_band, wb_band)
    first_order = BandOrder(ap_bands, ap_bands.ap_bands)
    first_order = first_order.shuffle()
    tmp = input("出演順を表示します．何かキーを押してください：")
    best_orders_top_bottom = CalcOrder.findBestOrderTopBottom(
        first_order, ap_bands)
    CalcOrder.print(best_orders_top_bottom)
    file1 = TextFile('top_bottom.txt', best_orders_top_bottom)
    file1.write()
    if best_orders_top_bottom[0].dup_sum != 0:
        print("トッパーとトリを指定して，重複しない組み合わせが見つかりませんでした")
    tmp = int(input("続けてお勧めの出演順を表示しますか : 1...表示する, 2...終了："))
    if tmp == 1:
        best_orders = CalcOrder.findBestOrder(first_order, ap_bands)
        CalcOrder.print(best_orders)
        file2 = TextFile('recommend.txt', best_orders)
        file2.write()
        if best_orders[0].dup_sum != 0:
            print("重複しない組み合わせが見つかりませんでした")
    elif tmp == 2:
        wait_exit()
    wait_exit()


class WorkBook:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
        except:
            print(f"{self.file_path} が正常に読み込めませんでした")
            wait_exit()
        self.ws = self.wb.active
        self.row_num = self.getRowNum(1)
        self.colum_num = self.getColumNum(1)
        self.row = [None for _ in range(self.row_num)]
        for i in range(self.row_num):
            self.row[i] = self.getRow(i + 1)
        self.colum = [None for _ in range(self.colum_num)]
        for i in range(self.colum_num):
            self.colum[i] = self.getColum(i + 1)
        self.wb.close()

   # colum列の行数を返す
    def getRowNum(self, colum):
        i = 0
        ln = self.ws.cell(i + 1, colum).value
        while not (ln == None):
            ln = self.ws.cell(i + 1, colum).value
            i += 1
        return i - 1

    # row行の列数を返す
    def getColumNum(self, row):
        i = 0
        ln = self.ws.cell(row, i + 1).value
        while not (ln == None):
            ln = self.ws.cell(row, i + 1).value
            i += 1
        return i - 1

    # colum列の内容を配列で返す
    def getColum(self, colum):
        row_num = self.getRowNum(colum)
        colum_data = [0 for _ in range(row_num)]
        for i in range(row_num):
            colum_data[i] = self.ws.cell(i + 1, colum).value
        return colum_data

    # row行の内容を配列で返す
    def getRow(self, row):
        colum_num = self.getColumNum(row)
        row_data = [0 for _ in range(colum_num)]
        for i in range(colum_num):
            row_data[i] = self.ws.cell(row, i + 1).value
        return row_data


class Bands:
    def __init__(self, wb_band) -> None:
        self.num = len(wb_band.row[1:])
        self.bands = [None for _ in range(self.num)]
        for i in range(self.num):
            self.bands[i] = Band(wb_band.row[i + 1])


class Band:
    def __init__(self, band_array) -> None:
        self.name = band_array[0]
        self.num = len(band_array[1:])
        self.member = band_array[1:]


class ApBands:
    def __init__(self, wb_ap_band, wb_band) -> None:
        self.num = len(wb_ap_band.row[1:])
        self.ap_bands = [None for _ in range(self.num)]
        self.top = wb_ap_band.row[1][3]
        self.bottom = wb_ap_band.row[1][4]
        for i in range(self.num):
            self.ap_bands[i] = ApBand(wb_ap_band.row[i + 1], wb_band)


class ApBand:
    def __init__(self, ap_band_array, wb_band) -> None:
        self.name = ap_band_array[0]
        self.time = ap_band_array[1]
        self.schedule = ap_band_array[2]
        self.index = self.isExist(wb_band)
        self.member = wb_band.row[self.index][1:]
        self.num = len(wb_band.row[1:])

    def isExist(self, wb_band):
        for i in range(len(wb_band.row[1:])):
            if self.name == wb_band.row[i + 1][0]:
                exist_index = i + 1
                return exist_index
            if i == len(wb_band.row[1:]) - 1:
                print(
                    f"{wb_band.file_path}にバンド：\"{self.name}\" が見つかりませんでした")
                wait_exit()


class BandOrder:  # バンド順を入力するとBandOrderクラスのインスタンスを生成
    def __init__(self, ap_bands, order) -> None:
        self.ap_bands = ap_bands
        self.order = copy.deepcopy(order)
        self.dup_list = self.getDupList()
        self.dup_sum = self.calcDupSum()
        self.continuous_max = self.howManyContinuousMax()

    def getDupList(self):
        dup_list = [[] for _ in range(len(self.ap_bands.ap_bands) - 1)]

        for i in range(len(dup_list)):
            dup = set(self.order[i].member) & set(self.order[i+1].member)
            dup = list(dup)
            dup_list[i] = dup
        return dup_list

    def calcDupSum(self):
        sum = 0
        for j in range(len(self.dup_list)):
            sum += len(self.dup_list[j])
        self.dup_sum = sum
        return self.dup_sum

    def print(self):
        time_sum = 0
        print("------------------------------バンド順-------------------------------")
        for i in range(len(self.order)):
            time_sum += self.order[i].time
            print(
                "{} : {} : {} 分 ({}分)".format(self.order[i].name, self.order[i].member, self.order[i].time, time_sum))
        print("---------------------------------------------------------------------")

    def exchange(self, i, j):
        new_order = copy.deepcopy(self.order)
        new_order[i], new_order[j] = new_order[j], new_order[i]
        new_order_instance = BandOrder(self.ap_bands, new_order)
        return new_order_instance

    def shuffle(self):
        new_order = copy.deepcopy(self.order)
        random.shuffle(new_order)
        new_order_instance = BandOrder(self.ap_bands, new_order)
        return new_order_instance  # バンド順を返す

    def shuffleTopBottom(self):  # top bottom 指定
        new_order = copy.deepcopy(self.order)
        tmp = new_order[1:len(new_order)]
        random.shuffle(tmp)
        new_order[1:len(new_order)] = tmp
        new_order_instance = BandOrder(self.ap_bands, new_order)
        return new_order_instance  # バンド順を返す

    def howManyContinuousMax(self):
        max = 0
        count = 0
        for i in range(len(self.dup_list) - 1):
            if set(self.dup_list[i]) & set(self.dup_list[i + 1]) != ():
                count += 1
            if count > max:
                max = count
        return count

    def isTheSame(self, another_order):
        for i in range(len(self.order)):
            if self.order[i].name != another_order.order[i].name:
                return False
        return True


class CalcOrder:
    @classmethod
    def findGoodOrder(self, order, ap_bands):  # 全ての順番を入れ替えて得られる，被る人数が最小のorderを返す
        best_dup_sum = 6*len(order.order)
        best_order = copy.deepcopy(order.order)
        for i in range(len(order.dup_list)):
            if not (order.dup_list[i] == []):
                for j in range(len(order.dup_list)):
                    if j == i or j == i + 1 or j == i + 2:
                        pass
                    order = order.exchange(i, j)
                    order.dup_list = order.getDupList()
                    order.dup_sum = order.calcDupSum()
                    if order.dup_sum < best_dup_sum:
                        best_dup_sum = order.dup_sum
                        best_order = copy.deepcopy(order.order)
                        break
                    order = order.exchange(i, j)
        best_order_instance = BandOrder(ap_bands, best_order)
        return best_order_instance

    @classmethod
    def findBestOrder(self, order, ap_bands):
        good_order_list = [order]
        print("算出中.....")
        for i in tqdm.tqdm(range(1000)):
            order = order.shuffle()
            order = CalcOrder.findGoodOrder(order, ap_bands)
            if order.dup_sum <= good_order_list[0].dup_sum and order.continuous_max <= good_order_list[0].continuous_max:
                if order.dup_sum < good_order_list[0].dup_sum or order.continuous_max < good_order_list[0].continuous_max:
                    good_order_list = []
                if self.isExistTheSameBand(good_order_list, order) == False:
                    good_order_list.append(order)
        return good_order_list

    @classmethod
    def print(self, good_order_list):
        for i in range(len(good_order_list)):
            good_order_list[i].print()
            print(good_order_list[i].dup_list)
        print("最適順番の候補", len(good_order_list), "個を表示しました")

    @classmethod
    def isExistTheSameBand(self, order_list, order):
        for i in range(len(order_list)):
            if order.isTheSame(order_list[i]) == True:
                return True
        return False

    @classmethod
    def findGoodOrderTopBottom(self, order, ap_bands):
        best_dup_sum = 6*len(order.order)
        i = 0
        while not (order.order[i].name == ap_bands.top):
            i += 1
        order.order[0], order.order[i] = order.order[i], order.order[0]

        i = 0
        while not (order.order[i].name == ap_bands.bottom):
            i += 1
        order.order[-1], order.order[i] = order.order[i], order.order[-1]
        best_order = copy.deepcopy(order.order)
        for i in range(1, len(order.dup_list)):
            if not (order.dup_list[i] == []):
                for j in range(1, len(order.dup_list)):
                    if j == i or j == i + 1 or j == i + 2:
                        pass
                    order = order.exchange(i, j)
                    order.dup_list = order.getDupList()
                    order.dup_sum = order.calcDupSum()
                    if order.dup_sum < best_dup_sum:
                        best_dup_sum = order.dup_sum
                        best_order = copy.deepcopy(order.order)
                        break
                    order = order.exchange(i, j)
        best_order_instance = BandOrder(ap_bands, best_order)
        return best_order_instance

    @classmethod
    def findBestOrderTopBottom(self, order, ap_bands):
        good_order_list = [order]
        print("算出中.....")
        for i in tqdm.tqdm(range(500)):
            order = order.shuffleTopBottom()
            order = CalcOrder.findGoodOrderTopBottom(order, ap_bands)
            if order.dup_sum <= good_order_list[0].dup_sum and order.continuous_max <= good_order_list[0].continuous_max:
                if order.dup_sum < good_order_list[0].dup_sum or order.continuous_max < good_order_list[0].continuous_max:
                    good_order_list = []
                if self.isExistTheSameBand(good_order_list, order) == False:
                    good_order_list.append(order)
        return good_order_list


class TextFile:
    def __init__(self, file_name, band_list):
        self.file_name = file_name
        self.band_list = band_list

    def write(self):
        f = open(self.file_name, 'w',  encoding='utf-8')
        f.write("------------------------------バンド順-------------------------------\n")
        f.write(f"候補バンド数：{len(self.band_list)}\n")
        for j in range(len(self.band_list)):
            time_sum = 0
            f.write(
                f"-------------------------------{j + 1}番目---------------------------------\n")
            for i in range(len(self.band_list[j].order)):
                time_sum += self.band_list[j].ap_bands.ap_bands[i].time
                f.write("{} : {} : {} 分 ({}分)\n".format(
                    self.band_list[j].order[i].name, self.band_list[j].order[i].member, self.band_list[j].order[i].time, time_sum))
            f.write(
                "---------------------------------------------------------------------\n")
            f.write(f"連続の人{self.band_list[j].dup_list}\n")
            f.write(
                "---------------------------------------------------------------------\n\n")
        f.close()
        print(f"{self.file_name}に結果を出力しました．")


def wait_exit():
    tmpo = input("何か押すと終了します：")
    exit()


if __name__ == "__main__":
    main()
